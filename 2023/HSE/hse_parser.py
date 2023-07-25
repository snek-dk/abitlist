import requests
from bs4 import BeautifulSoup
import os
import io
import openpyxl
import json


def get_json_form(main_form, personal_form):
    data_form = {
        "ВУЗ": "ВШЭ",
        "ОСНОВА_ОБУЧЕНИЯ": main_form["ОСНОВА_ОБУЧЕНИЯ"],
        "УРОВЕНЬ_ОБУЧЕНИЯ": main_form["УРОВЕНЬ_ОБУЧЕНИЯ"],
        "НАПРАВЛЕНИЕ": main_form["НАПРАВЛЕНИЕ"],
        "ОП": main_form["ОП"],
        "ФОРМА_ОБУЧЕНИЯ": main_form["ФОРМА_ОБУЧЕНИЯ"],
        "СНИЛС_УК": personal_form["СНИЛС_УК"],
        "ТИП_КОНКУРСА": personal_form["ТИП_КОНКУРСА"],
        "ПРИОРИТЕТ": personal_form["ПРИОРИТЕТ"],
        "СОГЛАСИЕ": personal_form["СОГЛАСИЕ"],
        "ПП": personal_form["ПП"],
        "ЕГЭ_С_ИД": personal_form["ЕГЭ_С_ИД"],
        "ЕГЭ": personal_form["ЕГЭ"],
        "ВИ_1": personal_form["ВИ_1"],
        "ВИ_2": personal_form["ВИ_2"],
        "ВИ_3": personal_form["ВИ_3"],
        "ВИ_4": personal_form["ВИ_4"],
        "ВИ_5": personal_form["ВИ_5"],
        "ВИ_6": personal_form["ВИ_6"],
        "ИД": personal_form["ИД"]
    }

    return data_form


def get_personal_form(data):
    subjects = data[8:-7]
    personal_form = {
        "СНИЛС_УК": data[0],
        "ТИП_КОНКУРСА": "БВИ" if data[1] == "Да" else (
            "Особая_квота" if data[2] == "Да" else (
                "Целевая_квота" if data[3] == "Да" else ("Отдельная_квота" if data[4] == "Да" else "ОК"))),
        "ПРИОРИТЕТ": data[6] if data[6] else None,
        "СОГЛАСИЕ": data[-4] if data[-4] else None,
        "ПП": data[-3] if data[-3] else None,
        "ЕГЭ_С_ИД": data[-6] if data[-6] else 0,
        "ЕГЭ": data[-6] if data[-6] else 0,
        "ВИ_1": None,
        "ВИ_2": None,
        "ВИ_3": None,
        "ВИ_4": None,
        "ВИ_5": None,
        "ВИ_6": None,
        "ИД": data[-7] if data[-7] else 0
    }

    for i in range(len(subjects)):
        personal_form[f"ВИ_{i + 1}"] = subjects[i] if subjects[i] is not None else 0

    return personal_form


main_page = BeautifulSoup(requests.get(url="https://ba.hse.ru/base2023").content.decode("utf-8"), "lxml")
universities = main_page.find_all("tbody")
universities = [university.find_all("tr")[1::] for university in universities]
hse_id = {
    0: "МСК",
    1: "НН",
    2: "ПЕРМЬ",
    3: "СПБ"
}
main_form = {
    "ВУЗ": "ВШЭ",
    "ОСНОВА_ОБУЧЕНИЯ": None,
    "УРОВЕНЬ_ОБУЧЕНИЯ": "Бакалавриат",
    "НАПРАВЛЕНИЕ": None,
    "ОП": None,
    "ФОРМА_ОБУЧЕНИЯ": None
}
for i in range(4):
    if not os.path.isdir(hse_id[i]):
        os.mkdir(hse_id[i])
for university_id in range(4):
    path = os.getcwd() + rf'\{hse_id[university_id]}'
    to_export = []
    main_form["ВУЗ"] += hse_id[university_id]
    for tr in universities[university_id]:
        if tr.a["href"] == "https://economics.hse.ru/EK_2023":
            continue
        print(tr.a["href"])
        xlsx = io.BytesIO(requests.get(url=tr.a["href"]).content)
        workbook = openpyxl.load_workbook(xlsx)
        worksheet = workbook.active
        needed_columns = [worksheet[15][column].value for column in range(0, worksheet.max_column)]
        program_name = ''
        program_code = ''
        form_of_education = ''
        for column in range(0, worksheet.max_column):
            program_name = program_name + worksheet[2][column].value if worksheet[2][
                                                                            column].value is not None else program_name
            program_code = program_code + worksheet[3][column].value if worksheet[3][
                                                                            column].value is not None else program_code
            form_of_education = form_of_education + worksheet[4][column].value if worksheet[4][
                                                                                      column].value is not None else form_of_education
        program_name = program_name.replace("Образовательная программа ", '', 1)
        program_code = program_code.replace("Направление ", '', 1)
        main_form["НАПРАВЛЕНИЕ"] = program_name
        main_form["ОП"] = program_code
        main_form["ФОРМА_ОБУЧЕНИЯ"] = form_of_education
        for row in worksheet.iter_rows(min_row=18, max_col=len(needed_columns), max_row=worksheet.max_row,
                                       min_col=2):
            parsed_data = []
            flag = False
            for cell in row:
                if flag:
                    parsed_data.append(cell.value)
                flag = not flag
            personal_info = get_personal_form(parsed_data)
            full_person_info = get_json_form(main_form, personal_info)
            if parsed_data[-5] not in ["Б", "К"]:
                full_person_info["ОСНОВА_ОБУЧЕНИЯ"] = "Бюджет"
                to_export.append(full_person_info)
                full_person_info["ОСНОВА_ОБУЧЕНИЯ"] = "Контракт"
                to_export.append(full_person_info)
            elif parsed_data[-5] == "Б":
                full_person_info["ОСНОВА_ОБУЧЕНИЯ"] = "Бюджет"
                to_export.append(full_person_info)
            else:
                full_person_info["ОСНОВА_ОБУЧЕНИЯ"] = "Контракт"
                to_export.append(full_person_info)
        xlsx.close()
    with open(os.getcwd() + rf'\{hse_id[university_id]}' + rf'\output.json', 'w') as f:
        json.dump(to_export, f, ensure_ascii=False, indent=4)
