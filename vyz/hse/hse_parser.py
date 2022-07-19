import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import json
from time import sleep


class HTMLTableParser:

    def parse_url(self, url):
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'lxml')
        return [(table, self.parse_html_table(table)) \
                for table in soup.find_all('table')]

    def parse_html_table(self, table):
        n_columns = 0
        n_rows = 0
        column_names = []

        # Find number of rows and columns
        # we also find the column titles if we can
        for row in table.find_all('tr'):

            # Determine the number of rows in the table
            td_tags = row.find_all('td')
            if len(td_tags) > 0:
                n_rows += 1
                if n_columns == 0:
                    # Set the number of columns for our table
                    n_columns = len(td_tags)

            # Handle column names if we find them
            th_tags = row.find_all('th')
            if len(th_tags) > 0 and len(column_names) == 0:
                for th in th_tags:
                    column_names.append(th.get_text())

        # Safeguard on Column Titles
        if len(column_names) > 0 and len(column_names) != n_columns:
            raise Exception("Column titles do not match the number of columns")

        columns = column_names if len(column_names) > 0 else range(0, n_columns)
        df = pd.DataFrame(columns=columns,
                          index=range(0, n_rows))
        row_marker = 0
        for row in table.find_all('tr'):
            column_marker = 0
            columns = row.find_all('td')
            for column in columns:
                df.iat[row_marker, column_marker] = column.get_text()
                column_marker += 1
            if len(columns) > 0:
                row_marker += 1

        # Convert to float if possible
        for col in df:
            try:
                df[col] = df[col].astype(float)
            except ValueError:
                pass

        return df


hp = HTMLTableParser()
table = hp.parse_url("https://ba.hse.ru/stat2022")[0][1]
# чтение всех направлений
for i in range(1,len(table)):
    name = table[0][i]
    link = table[1][i]
    name_file = link.split("/")[-1]
    # скачивание файла
    response = requests.get(link)
    open(name_file, "wb").write(response.content)
    wb = load_workbook(str(name_file))
    sheet = wb.active
    ed_program = "_".join(sheet.cell(row=2, column=3).value.split()[2:]).replace('"','')
    field_of_study = "_".join(sheet.cell(row=3, column=3).value.split()[1:])
    # спросить с какой буквы форма обучения Б или М
    form_of_education = sheet.cell(row=4, column=3).value.split()[0]
    col = 18
    collected_data = dict()
    count = 1
    while True:
        snils = sheet.cell(row=col, column=3).value
        if snils == None: break
        contest = 'По результатам ВИ' if sheet.cell(row=col, column=5).value == "Нет" else "БВИ"
        all_sum = sheet.cell(row=col, column=21).value if sheet.cell(row=col, column=21).value is not None else 0
        sum_without_id = (int(sheet.cell(row=col, column=21).value) if sheet.cell(row=col,
                                                                                  column=21).value is not None else 0) - (
                             int(sheet.cell(row=col, column=19).value) if sheet.cell(row=col,
                                                                                     column=19).value is not None else 0)
        math = sheet.cell(row=col, column=13).value if sheet.cell(row=col, column=13).value is not None else 0
        inf = sheet.cell(row=col, column=15).value if sheet.cell(row=col, column=15).value is not None else 0
        rus = sheet.cell(row=col, column=17).value if sheet.cell(row=col, column=17).value is not None else 0
        id = int(sheet.cell(row=col, column=19).value) if sheet.cell(row=col, column=19).value is not None else 0
        agreement = sheet.cell(row=col, column=27).value
        original = sheet.cell(row=col, column=25).value
        ed_principle = sheet.cell(row=col, column=23).value
        preemptive_right = sheet.cell(row=col, column=29).value
        result = {'ВУЗ': 'НИУ ВШЭ',
                  'Направление': field_of_study,
                  'ОП': ed_program,
                  'Форма_обучения': form_of_education,
                  'Основа_обучения': 'госбюджетная',
                  snils: {'Конкурс': contest,
                                     'Приоритет': None,
                                     'СУММА': int(all_sum),
                                     'СУММА_БЕЗ_ИД': int(sum_without_id),
                                     'ВИ1': int(math),
                                     'ВИ2': int(inf),
                                     'ВИ3': int(rus),
                                     'ИД': int(id),
                                     'Согласие':agreement,
                                     'Оригинал': original,
                                     'ПП': preemptive_right}}

        if ed_principle == "Б":
            result['Основа_обучения'] = 'госбюджетная'
            collected_data[count] = result.copy()
            count += 1
        if ed_principle == "К; Б" or ed_principle == "Б; К":
            result['Основа_обучения'] = 'госбюджетная'
            collected_data[count] = result.copy()
            count += 1
            result['Основа_обучения'] = 'контракт'
            collected_data[count] = result.copy()
            count += 1
        if ed_principle == "К":
            result['Основа_обучения'] = 'контракт'
            collected_data[count] = result.copy()
            count += 1
        col += 1


    with open(r"hse" + str(name_file[:-5])+ ".json", 'w') as fp:
        json.dump(collected_data, fp, indent=4)
    wb.close()
    sleep(10)