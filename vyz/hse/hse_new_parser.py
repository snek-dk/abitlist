import openpyxl


def insert_into_dict(privileges_w_snils, subjects, other_data, main_data):
    abit_id = privileges_w_snils[0]
    entrance_types = ['БВИ', 'ОП', 'ЦП', 'СК', 'ОК']
    if privileges_w_snils[1:].count('Да') > 0:
        entrance = entrance_types[privileges_w_snils.index('Да')]
    else:
        entrance = entrance_types[-1]
    while len(subjects) < 5:
        subjects.append(None)
    city, field_of_study, ed_program = main_data[0], main_data[1], main_data[2]
    try:
        bez_id = str(int(other_data[1]) - int(other_data[0]))
    except:
        bez_id = other_data[1]
    entrant = {
        'ВУЗ': 'ВШЭ' + ' ' + city,
        'Направление': field_of_study,
        'ОП': ed_program,
        'Форма_обучения': "очная",
        'Основа_обучения': other_data[2],
        'СНИЛС_УК': abit_id,
        'Конкурс': entrance,
        'СУММА': other_data[1] if other_data[1] else "0",
        'СУММА_БЕЗ_ИД': bez_id if bez_id else "0",
        'ВИ_1': subjects[0] if subjects[0] else "0",
        'ВИ_2': subjects[1] if subjects[1] else "0",
        'ВИ_3': subjects[2] if subjects[2] else "0",
        'ВИ_4': subjects[3] if subjects[3] else "0",
        'ВИ_5': subjects[4] if subjects[4] else "0",
        'ИД': other_data[0] if other_data[0] else "0",
        'Согласие': other_data[-1],
        'Оригинал': other_data[-2]
    }
    return entrant


path = r"C:\Users\dmitr\PycharmProjects\abitlist\vyz\hse\tables\moscow\Информатика_и_вычислительная_техника.xlsx"
book = openpyxl.open(path, read_only=True, data_only=True)
sheet = book.active
needed_columns = set(sheet[15][column].value for column in range(0, sheet.max_column))
main_data = [path.split('\\')[-2], path.split('\\')[-1][:-5].replace('_', ' '),
             path.split('\\')[-1][:-5].replace('_', ' ')]
to_save = []
k = 0

for row in sheet.iter_rows(min_row=18, max_col=len(needed_columns) * 2, max_row=sheet.max_row,
                           min_col=2):  # range(18, sheet.max_row):
    parsed_data = []
    flag = False
    for cell in row:  # range(0, (len(needed_columns) - 1) * 2, 2):
        if flag:
            parsed_data.append(cell.value)
        flag = not flag
    # print(parsed_data)
    privileges_w_snils = parsed_data[:5]
    # print(privileges_w_snils)
    other_data = parsed_data[-9:-4]
    # print(other_data)
    subjects = parsed_data[6:-8]
    # print(subjects)
    if other_data[2] is None:
        # print(parsed_data)
        # print(row)
        continue
    if len(other_data[2]) > 1:
        other_data[2] = "Госбюджет"
        to_save.append(insert_into_dict(privileges_w_snils=privileges_w_snils, subjects=subjects, other_data=other_data,
                                        main_data=main_data))
        other_data[2] = "Контракт"
        to_save.append(insert_into_dict(privileges_w_snils=privileges_w_snils, subjects=subjects, other_data=other_data,
                                        main_data=main_data))
    elif other_data[2] == "Б":
        other_data[2] = "Госбюджет"
        to_save.append(insert_into_dict(privileges_w_snils=privileges_w_snils, subjects=subjects, other_data=other_data,
                                        main_data=main_data))
    else:
        other_data[2] = "Контракт"
        to_save.append(insert_into_dict(privileges_w_snils=privileges_w_snils, subjects=subjects, other_data=other_data,
                                        main_data=main_data))
print(to_save)
